#!/usr/bin/Python3.9

import json
import openpyxl
import os
from JsonToExcel import *
import platform

if platform.system() == 'Darwin':
    ACCOUNT_HISTORY_PATH = "{0}/../data/accountHistory.json".format(
        os.path.dirname(__file__))
elif platform.system() == 'Windows':
    ACCOUNT_HISTORY_PATH = "{0}\..\data\accountHistory.json".format(
        os.path.dirname(__file__))
else:
    ACCOUNT_HISTORY_PATH = "{0}/../data/accountHistory.json".format(
        os.path.dirname(__file__))


def updateMemory(filePATH):
    with open(ACCOUNT_HISTORY_PATH, "r") as f:
        accountHistory = json.load(f)
    workbook = openpyxl.load_workbook(filePATH)
    worksheet = workbook.active
    descriptions = []
    comptes = []
    taxRegime = []
    for i in range(worksheet.max_row-3):
        descriptions.append(worksheet.cell(
            i+2, COLUMNS["DESCRIPTION"]+1).value)
        comptes.append(worksheet.cell(i+2, COLUMNS["COMPTE"]+1).value)
        taxRegime.append(worksheet.cell(
            i+2, COLUMNS["REGIME DE TAXE"]+1).value)
    for i in range(len(descriptions)):
        if comptes[i] is not None:
            accountHistory[descriptions[i]] = {
                "COMPTE": comptes[i],
                "TAX REGIME": taxRegime[i]
            }
    with open(ACCOUNT_HISTORY_PATH, "w+") as f:
        json.dump(accountHistory, f)
